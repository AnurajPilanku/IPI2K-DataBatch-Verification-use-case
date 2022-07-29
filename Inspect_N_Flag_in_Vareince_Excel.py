#AnurajPilanku
#Check for any "N"in varience.xlsx
import openpyxl
import sys
from datetime import date, timedelta

tday=date.today()
yday=tday-timedelta()

varpath=sys.argv[1]
varwb=openpyxl.load_workbook(varpath)
varsh=varwb.worksheets[0]

dt=[]
for i in range(2,varsh.max_row+1):
    if varsh.cell(column=6,row=i).value not in [None,""," "]:
        if str(varsh.cell(column=6,row=i).value)[:10] in [yday,str(yday)]:
            dt.append(varsh.cell(column=4,row=i).value)
if "N" in dt:
    print("Presence of N")
else:
    print("Absence of N")
varwb.close()
