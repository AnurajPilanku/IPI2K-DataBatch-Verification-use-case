#Anuraj Pilanku
#Comparing todays data of N flag with yesterdays notepad data

import openpyxl
import sys
from datetime import date, timedelta
import calendar

tday=date.today()
day=calendar.day_name[tday.weekday()]
if day in ["Tuesday"]:
    yday = tday - timedelta(4)
else:
    yday = tday - timedelta(1)


todaysNflagpath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\file_with_required_data\blank_removed.xlsx"#sys.argv[1]#blank_remove.xlsx--current flow execution xlsx which contains N flag data-
yesterdaysnotepadpath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\backupdata\Headerbackup.xlsx"#sys.argv[2]#heasderbackup.xlsx

todaysNflagpathwb=openpyxl.load_workbook(todaysNflagpath)
todaysNflagpathsh=todaysNflagpathwb.worksheets[0]

yesterdaysnotepadwb=openpyxl.load_workbook(yesterdaysnotepadpath)
yesterdaysnotepadsh=yesterdaysnotepadwb.worksheets[0]
#Creating dictionory of yesterdays notepad data
yesterdaysnotepaddict=dict()
for r in range(2,yesterdaysnotepadsh.max_row+1):
    if yesterdaysnotepadsh["E"+str(r)].value not in [None,""," "]:
        if str(yesterdaysnotepadsh["E"+str(r)].value)[:10] in [yday,str(yday)]:#[2,"2"]:
            if yesterdaysnotepadsh["A"+str(r)].value not in [None,""," "]:
                if yesterdaysnotepadsh["C" + str(r)].value not in [None, "", " "]:
                    yesterdaysnotepaddict[str(yesterdaysnotepadsh["A"+str(r)].value).strip()+str(yesterdaysnotepadsh["C"+str(r)].value).strip()]=str(yesterdaysnotepadsh["C"+str(r)].value).strip()
#print(yesterdaysnotepaddict)
for r in range(2,todaysNflagpathsh.max_row+1):
    if todaysNflagpathsh["D"+str(r)].value=="N":
        if todaysNflagpathsh["B"+str(r)].value not in [None,""," "]:
            if todaysNflagpathsh["C" + str(r)].value not in [None, "", " "]:
                if str(todaysNflagpathsh["B"+str(r)].value).strip()+str(todaysNflagpathsh["C"+str(r)].value).strip() in yesterdaysnotepaddict:
                    todaysNflagpathsh["E" + str(r)].value=yesterdaysnotepaddict[str(todaysNflagpathsh["B"+str(r)].value).strip()+str(todaysNflagpathsh["C"+str(r)].value).strip()]
                    todaysNflagpathsh["D" + str(r)].value="Y"
flaglist=list()
for r in range(2,todaysNflagpathsh.max_row+1):
    flaglist.append(todaysNflagpathsh["D" + str(r)].value)
if "N" in flaglist:
    print("failure")
else:
    print("success")
todaysNflagpathwb.save(todaysNflagpath)


