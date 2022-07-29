#Anuraj Pilanku
#copy notepad data to headerbackup.xlsx

import openpyxl
import sys
from datetime import date, timedelta
import string

tday=date.today()
yday=tday-timedelta(1)

todaycphpath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\file_with_required_data\req_columns.xlsx"#sys.argv[1]
headerbackuppath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\backupdata\Headerbackup.xlsx"#sys.argv[2]

todaycphwb=openpyxl.load_workbook(todaycphpath)
todaycphsh=todaycphwb.worksheets[1]

headerbackupwb=openpyxl.load_workbook(headerbackuppath)
headerbackupsh=headerbackupwb.worksheets[0]
final=headerbackupsh.max_row
for i in range(1,todaycphsh.max_column+1):
    for r in range(2,todaycphsh.max_row + 1):
        #headerbackupsh[string.ascii_uppercase[i-1]+str(r)].value=todaycphsh[string.ascii_uppercase[i-1]+str(r)].value
        headerbackupsh.cell(column=i,row=r+final).value=todaycphsh.cell(column=i,row=r).value
for r in range(2,todaycphsh.max_row + 1):
    headerbackupsh['E'+str(r+final)].value=tday#1
headerbackupwb.save(headerbackuppath)
print("success")