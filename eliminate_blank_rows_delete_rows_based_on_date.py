'''
Created by     : AnurajPilanku
Code utility   : Delete rows based on Date & Remove Blank rows from Excel file and shift rows up
Use case        : IPI2K Use case
'''
import openpyxl
import sys
from datetime import date,timedelta
import calendar
import os

tday=date.today()
yday=tday-timedelta(5)

day=calendar.day_name[tday.weekday()]
if day in ["Tuseday"]:
    yday = tday - timedelta(5)
tobeRenamedPath=sys.argv[1]
renamedPath=sys.argv[2]

os.rename(tobeRenamedPath,renamedPath)#(r"C:\Users\2040664\anuraj\ipi2k\testjh.xlsx",r"C:\Users\2040664\anuraj\ipi2k\renamedzzzz.xlsx")

oldwb=openpyxl.load_workbook(renamedPath)
oldsh=oldwb.worksheets[0]

newwb=openpyxl.Workbook()
newsh=newwb.active

for col in range(1,oldsh.max_column+1):
    newsh.cell(column=col,row=1).value=oldsh.cell(column=col,row=1).value

#for col in range(1,oldsh.max_column+1):
li1=[]
li2=[]
li3=[]
li4=[]
li5=[]
for row in range(2,oldsh.max_row+1):
    if oldsh.cell(column=1,row=row).value and oldsh.cell(column=2,row=row).value and oldsh.cell(column=3,row=row).value not in [None,""," "]:
        if oldsh.cell(column=5,row=row).value.date() >yday:
            li1.append(oldsh.cell(column=1,row=row).value)
            li2.append(oldsh.cell(column=2, row=row).value)
            li3.append(oldsh.cell(column=3, row=row).value)
            li4.append(oldsh.cell(column=4, row=row).value)
            li5.append(oldsh.cell(column=5, row=row).value)
for row in range(0,len(li1)):
    newsh.cell(column=1,row=row+2).value=li1[row]
    newsh.cell(column=2, row=row + 2).value = li2[row]
    newsh.cell(column=3, row=row + 2).value = li3[row]
    newsh.cell(column=4, row=row + 2).value = li4[row]
    newsh.cell(column=5, row=row + 2).value = li5[row].date()#[:10]
newwb.save(tobeRenamedPath)
os.remove(renamedPath)
print("success")

#python delrow.py C:\Users\2040664\anuraj\ipi2k\Headerbackup.xlsx C:\Users\2040664\anuraj\ipi2k\Headerbackupdelfor.xlsx




