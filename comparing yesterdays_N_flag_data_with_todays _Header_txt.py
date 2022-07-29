#AnurajPilanku
# yesterdayNflagdata.xlsx
#varience.xlsx
#Checking whether yesterdays data corresponding to N flag is matching with todays Header.TXT data
import openpyxl
import sys
from datetime import date, timedelta
import calendar

tday=date.today()
day=calendar.day_name[tday.weekday()]
if day in ["Tuesday"]:
    yday = tday - timedelta(4)
else:
    yday = tday - timedelta(1)


todaynotepadpath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\file_with_required_data\req_columns.xlsx"#sys.argv[1]
yesterdayNflagdatapath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\backupdata\varience.xlsx"#sys.argv[2]

todaynotepadwb=openpyxl.load_workbook(todaynotepadpath)
todaynotepadsh=todaynotepadwb.worksheets[1]

yesterdayNflagdatawb=openpyxl.load_workbook(yesterdayNflagdatapath)
yesterdayNflagdatash=yesterdayNflagdatawb.worksheets[0]

#step1:adding one with age
#for r in range(2,yesterdayNflagdatash.max_row+1):
    #if yesterdayNflagdatash['F'+str(r)].value not in [None,"NULL"," ",'']:
        #yesterdayNflagdatash['F'+str(r)].value=int(yesterdayNflagdatash['F'+str(r)].value)+1

#step2 :creating dict of todays header.txt
todays_notepad_data_dic=dict()
for i in range(2,todaynotepadsh.max_row+1):
    if todaynotepadsh['A'+str(i)].value not in [None,""," "]:
        if todaynotepadsh['C' + str(i)].value not in [None, "", " "]:
            todays_notepad_data_dic[str(todaynotepadsh['A'+str(i)].value).strip()+str(todaynotepadsh['C'+str(i)].value).strip()]=str(todaynotepadsh['C'+str(i)].value).strip()

#step3 checking whether yesterdays data corresponding to flag n which is stored in varience.xlsx is present in dictionory(todays notepad data)
for r in range(2,yesterdayNflagdatash.max_row+1):
    if yesterdayNflagdatash['F'+str(r)].value not in [None,""," "]:
        if str(yesterdayNflagdatash['F'+str(r)].value)[:10] in [yday,str(yday)]:#[2,"2"]:
            if str(yesterdayNflagdatash['B'+str(r)].value).strip()+str(yesterdayNflagdatash['C'+str(r)].value).strip() in todays_notepad_data_dic:
                yesterdayNflagdatash['E' + str(r)].value=todays_notepad_data_dic[str(yesterdayNflagdatash['B'+str(r)].value).strip()+str(yesterdayNflagdatash['C'+str(r)].value).strip()]
                yesterdayNflagdatash['D' + str(r)].value="Y"
            else:
                yesterdayNflagdatash['D' + str(r)].value="N"

#step4:Determining success or failure: by checking whether there is any "N" in the list which contains Flag values of yesterdays data in varience .xlsx
flaglist=list()
for r in range(2,yesterdayNflagdatash.max_row+1):
    if yesterdayNflagdatash['F' + str(r)].value not in [None, "", " "]:
        if str(yesterdayNflagdatash['F' + str(r)].value)[:10] in [yday, str(yday)]:
            flaglist.append(yesterdayNflagdatash['D' + str(r)].value)
if "N" in flaglist:
    print("failure")
else:
    print("success")
yesterdayNflagdatawb.save(yesterdayNflagdatapath)