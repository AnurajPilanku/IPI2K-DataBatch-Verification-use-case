#AnurajPilanku
#IPI2K
#Get data which corresponds to "N" in blankremoved excel workbook

import openpyxl
import sys
from datetime import date, timedelta

tday=date.today()
yday=tday-timedelta(1)

blankremovedpath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\file_with_required_data\blank_removed.xlsx"#sys.argv[1]
varpath=r"\\acprd01\E\3M_CAC\IPI2K_DBV\backupdata\varience.xlsx"#sys.argv[2]

blankremwb=openpyxl.load_workbook(blankremovedpath)
blankremsh=blankremwb.worksheets[0]

varwb=openpyxl.load_workbook(varpath)
varsh=varwb.worksheets[0]
final=varsh.max_row
#print(final)

for i in range(2,blankremsh.max_row+1):
    if blankremsh.cell(column=4,row=i).value=="N":
        varsh.cell(column=1, row=i+final-2).value=blankremsh.cell(column=1,row=i).value
        varsh.cell(column=2, row=i+final-2).value = blankremsh.cell(column=2, row=i).value
        varsh.cell(column=3, row=i+final-2).value = blankremsh.cell(column=3, row=i).value
        varsh.cell(column=4, row=i+final-2).value = blankremsh.cell(column=4, row=i).value
        varsh.cell(column=5, row=i+final-2).value = blankremsh.cell(column=5, row=i).value
        varsh.cell(column=6, row=i+final-2).value= tday
varwb.save(varpath)
print("success")



