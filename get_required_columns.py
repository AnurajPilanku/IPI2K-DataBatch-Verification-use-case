'''

Created By   : Anuraj Pilanku
Use case     : IPI2K Daily Batch Verification
Code Utility  : get required columns from xls to xlsx

'''

import pandas as pd
import sys
import openpyxl
import time
#import xlsxwriter
filepath=sys.argv[1]
outfilepath=sys.argv[2]
ipidt=pd.read_excel(filepath,sheet_name = 0,engine='xlrd',dtype=str, index_col = 1,encoding='latin1',na_filter=False)#encoding_override='CORRECT_ENCODING')


#getting sheet names in an excel
sheetnames=ipidt.keys()
columnnames=list(ipidt.columns.values)
req_data=ipidt[[columnnames[0],columnnames[2],columnnames[3]]]
#req_data=ipidt.iloc[:,[0,3,4]]
req_data.to_excel(outfilepath,sheet_name=sheetnames[0],index=False)


time.sleep(3)
wb=openpyxl.load_workbook(outfilepath)
ws=wb.worksheets[0]
for i in range(2,ws.max_row+1):
    if ws.cell(column=3,row=i).value not in [None,""," "]:
        ws.cell(column=3, row=i).value=str(ws.cell(column=3,row=i).value).lstrip("0")

        #if str(ws.cell(column=3,row=i).value)[:1]=="0":
            #ws.cell(column=3, row=i).value=str(ws.cell(column=3,row=i).value)[1:]
wb.save(outfilepath)
print("success")
