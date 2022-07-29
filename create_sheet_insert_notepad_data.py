'''
Created by     : AnurajPilanku
Code utility   : Transfering Data from Notepad to XLSX file
Usecase        : IPI2K Use case
'''

# import libraries
import time
import os
import pandas as pd
import sys
import openpyxl

notepadOnePath=sys.argv[2]
notepadTwoPath=sys.argv[3]
excelpath=sys.argv[1]

#dealing missing values in notepad
def npad(npth):
    npadpath =npth #r"C:\Users\2040664\anuraj\promise\Header_extract.TXT"  # sys.argv[1]
    with open(npadpath, mode='r', encoding='utf-8') as f:
        data = ','.join(f.readlines()).replace(",", "")
    flist = list()
    list_data = data.split("\n")
    for i in list_data:
        cv = i.split()
        if len(cv) not in [4,0]:
            if cv[1].isdigit():
                cv.insert(1, "null")
                jnd = "\t".join(cv)
                flist.append(jnd + "\n")
        else:
            jnd = "\t".join(cv)
            flist.append(jnd + "\n")
    #print(flist)

    with open(npadpath, mode='w', encoding='utf-8') as f:
        [f.write(i) for i in flist]
    f.close()
if os.stat(notepadOnePath).st_size!=0:
    npad(notepadOnePath)
if os.stat(notepadTwoPath).st_size!=0:
    npad(notepadTwoPath)
time.sleep(3)

hd = 'Site_Code,cde,PH2_Btach_Number,cd'.split(",")

data = pd.read_csv(notepadOnePath, delim_whitespace=True, header=None, names=hd, index_col=False, na_filter=False)
data2 = pd.read_csv(notepadTwoPath, delim_whitespace=True, header=None, names=hd, index_col=False, na_filter=False)
# combineddata=pd.concat([data,data2],axis=1,ignore_index=True)
combineddata = data.append(data2)

wb=openpyxl.load_workbook(excelpath)
wb.create_sheet('CPS BATCH',index=1)
sheet = wb.worksheets[1]
s=wb.worksheets[0]
for r in range(0, len(combineddata)):
    for c in range(0, len(combineddata.columns)):
        sheet.cell(column=c + 1, row=r + 2).value = list(combineddata.iloc[:, c])[r]
for i in range(0, len(hd)):
    sheet.cell(column=i + 1, row=1).value = hd[i]
#print("successfully transfered notepad Data")
# combineddata.to_excel(path2, index = False)

s['D1'].value='Confirmed'
s['E1'].value='CPS Batch'


dic=dict()
li=list()
for i in range(2,sheet.max_row+1):
    li.append(sheet.cell(column=1,row=i).value)
    dic[str(sheet.cell(column=1,row=i).value)+str(sheet.cell(column=3,row=i).value)]=sheet.cell(column=3,row=i).value
absentValues=[]
secli=list()
for i in range(2,s.max_row+1):
    secli.append(s.cell(column=2,row=i).value)
    x=str(s.cell(column=2,row=i).value)+str(s.cell(column=3,row=i).value)
    if x in dic.keys():
        s.cell(column=5, row=i).value=str(dic[x])#qaqa
    #else:
       # if s.cell(column=3,row=i).value not in [None," "]:
            #absentValues.append(str(s.cell(column=2,row=i).value)+":"+str(s.cell(column=3,row=i).value))
#Filling the non common cells
sheet2ValNotInsheet1=dict()
phbatchnumsheet1=list()
for i in range(2,s.max_row+1):
    if s.cell(column=3, row=i).value not in [None,'#NA'," "]:
        phbatchnumsheet1.append(str(s.cell(column=3, row=i).value))

for i in range(2,sheet.max_row+1):
    if str(sheet.cell(column=3, row=i).value) not in phbatchnumsheet1:
        sheet2ValNotInsheet1[sheet.cell(column=1, row=i).value]=str(sheet.cell(column=3, row=i).value)
for i in range(2,s.max_row+1):
    if s.cell(column=3,row=i).value not in [None,"#NA"," "]:
        if s.cell(column=5, row=i).value in [None,"#NA", " "]:
            if s.cell(column=2, row=i).value in sheet2ValNotInsheet1.keys():
                s.cell(column=5, row=i).value=sheet2ValNotInsheet1[s.cell(column=2, row=i).value]

wb.save(excelpath)

print("success")

